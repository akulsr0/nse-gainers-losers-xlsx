from nsetools import Nse
from openpyxl import Workbook

nse = Nse()
wb = Workbook()


def get_top_gainers():
    top_gainers = nse.get_top_gainers()
    top_gainers_sheet = wb.create_sheet('Top Gainers')
    top_gainers_sheet.append(['No.', 'Symbol', 'Previous Close', 'Last Traded Price', 'Gain', 'Gain %'])
    for i, tg in enumerate(top_gainers):
        symbol = tg['symbol']
        prev = tg['previousPrice']
        ltp = tg['ltp']
        profit = round(ltp - prev, 2)
        profit_per = str(round(profit*100/prev, 2))+'%'
        top_gainers_sheet.append([i+1, symbol, prev, ltp, profit, profit_per])


def get_top_losers():
    top_losers = nse.get_top_losers()
    top_losers_sheet = wb.create_sheet('Top Losers')
    top_losers_sheet.append(['No.', 'Symbol', 'Previous Close', 'Last Traded Price', 'Loss', 'Loss %'])
    for i, tg in enumerate(top_losers):
        symbol = tg['symbol']
        prev = tg['previousPrice']
        ltp = tg['ltp']
        loss = round(prev - ltp, 2)
        loss_per = str(round(loss*100/ltp, 2))+'%'
        top_losers_sheet.append([i+1, symbol, prev, ltp, loss, loss_per])


get_top_gainers()
get_top_losers()
wb.save('output.xlsx')
